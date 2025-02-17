# !/usr/bin/env python
# coding= utf-8
"""
@Project         : Saudi
@File            : compare-csv.py
@Author          : yum <richardminyu@foxmail.com>
@Date            : 2024/05/08 15:43
@Description     : 比较两个 .xlsx 文件的差异，在第二个参与比较的文件中显示不同。
                   COMMAND: py compare-excel.py file1 file2

TODO: 局限，只适合行列一致，同单元格的数据检查
"""
# TODO:Completed
import time
import math
import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill


def compare_excel(file_fir, file_sec):
    if not os.path.exists(file_fir):
        print("请输入要比较的文件")
        return

    if not os.path.exists(file_sec):
        print("请输入要比较的文件")
        return

    print("------ comparing... ------")

    # 读取 excel 文件
    xlsx_aa = openpyxl.load_workbook(file_fir, data_only=True)
    xlsx_bb = openpyxl.load_workbook(file_sec, data_only=True)

    # 获取活动表（默认表）
    xlsx_a = xlsx_aa.active
    xlsx_b = xlsx_bb.active

    # 获取行列数
    # row_max = xlsx_a.max_row if xlsx_a.max_row > xlsx_b.max_row else xlsx_b.max_row
    # col_max = xlsx_a.max_column if xlsx_a.max_row > xlsx_b.max_column else xlsx_b.max_column
    row_max = 3971
    col_max = 19
    # print(row_max, col_max)

    # 设置单元格样式
    diff_cell_font_1 = Font(name='等线', size=20, italic=True, color='FF0000', bold=True)
    # diff_cell_font_2 = Font(name='等线', size=14, italic=True, color='FF0000', bold=True)
    diff_cell_fill = PatternFill(fill_type=None, start_color='98FB98', end_color='00FF7F')

    # 映射列
    chara_list = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S',
                  'T',
                  'U', 'V', 'W', 'X', 'Y', 'Z']

    # 便利所有单元格
    for i in range(12, row_max + 1):
        for j in range(1, col_max + 1):
            # 获取单元格数据
            a_val = xlsx_a.cell(row=i, column=j).value
            b_val = xlsx_b.cell(row=i, column=j).value
            # print('0--0', i, j)
            # print('-00-', a_val, b_val)
            # print()
            # print('00', '[', i, j, ']', b_val)

            if a_val is not None and b_val is not None and a_val != b_val:
                # 都非空值，且不等
                # print('-----------------------------------------------------------')
                # print('-11-', xlsx_a['' + str(chara_list[j]) + str(i)], xlsx_a['' + str(chara_list[j]) + str(i)].value)
                # print('-22-', xlsx_b['' + str(chara_list[j]) + str(i)], xlsx_b['' + str(chara_list[j]) + str(i)].value)
                # print('-----------------------------------------------------------')
                xlsx_b['' + str(chara_list[j]) + str(i)].font = diff_cell_font_1
                xlsx_b['' + str(chara_list[j]) + str(i)].fill = diff_cell_fill
            elif a_val is None and b_val is not None:
                # 标准值为空，比较值不为空
                xlsx_b['' + str(chara_list[j]) + str(i)].font = diff_cell_font_1
                xlsx_b['' + str(chara_list[j]) + str(i)].fill = diff_cell_fill
            elif a_val is not None and b_val is None:
                # 标准值不为空，比较值为空
                xlsx_b['' + str(chara_list[j]) + str(i)].value = 'None'
                xlsx_b['' + str(chara_list[j]) + str(i)].font = diff_cell_font_1
                xlsx_b['' + str(chara_list[j]) + str(i)].fill = diff_cell_fill

    xlsx_bb.save(f"./diff_{os.path.basename(file_fir)}_{os.path.basename(file_sec)}_{math.ceil(time.time())}.xlsx")
    print("------ success ------")


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("请输入合理的参数，如：py compare-excel.py file-1 file-2")
    else:
        # print(sys.argv[1], sys.argv[2], sys.argv)
        compare_excel(sys.argv[1], sys.argv[2])
